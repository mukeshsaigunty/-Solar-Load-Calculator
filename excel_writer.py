from openpyxl import load_workbook
import os
from datetime import datetime, timedelta

def fill_excel(data, template_path, output_path):
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"{template_path} not found")

    wb = load_workbook(template_path)
    ws = wb.active

    # Fill consumer details (LEFT COLUMN - D)
    if data.get("consumer_name"):
        ws["D1"] = data["consumer_name"]
    
    if data.get("consumer_no"):
        ws["D2"] = data["consumer_no"]
    
    if data.get("fixed_charges"):
        ws["D3"] = data["fixed_charges"]
    
    if data.get("load"):
        ws["D4"] = f"{data['load']}KW"
    
    if data.get("connection_type"):
        ws["D5"] = data["connection_type"]
    
    # Fill monthly data (starting from row 9)
    monthly_data = data.get("monthly_data", [])
    
    if monthly_data:
        base_date = datetime(2025, 2, 1)  # Feb 2025
        
        for idx, units in enumerate(monthly_data[:12]):
            row = 9 + idx
            
            # Month column (C)
            month_date = base_date + timedelta(days=30 * idx)
            ws[f"C{row}"] = month_date
            
            # Units column (D)
            ws[f"D{row}"] = units
            
            # Bill amount column (E) - use provided amount or calculate
            if idx == len(monthly_data) - 1 and data.get("current_amount"):
                ws[f"E{row}"] = data["current_amount"]
            else:
                ws[f"E{row}"] = ""
    
    # The template has formulas that will auto-calculate:
    # - Average consumption (D22)
    # - Recommended kW (D23)
    # - Solar panels needed (D24)
    # - Solar capacity (D25)
    # - Number of panels (D26)
    
    wb.save(output_path)